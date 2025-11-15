import requests
import pandas as pd
import os
import re
import time
from openpyxl import load_workbook
import json


# --- CẤU HÌNH ---
# API endpoint của trang web
api_url = r'https://ss.shanghai.gov.cn/manda-app/api/app/search/v1/vgyfvp/search'

# Các biến điều khiển quá trình crawl
request_delay = 0.5  # Giây, thời gian chờ giữa các request
retry_attempts = 3   # Số lần thử lại nếu request thất bại
retry_delay = 5      # Giây, thời gian chờ trước khi thử lại

# Tên các file output và checkpoint
checkpoint_file = 'checkpoint.log'
txt_filename = 'crawled_titles.txt'
excel_filename = '模版.xlsx'
sheet_name_to_update = 'shanghai'

# Headers để giả mạo là một trình duyệt hợp lệ
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 QuarkPC/4.6.5.580',
    'Content-Type': 'application/json',
    'Accept': 'application/json, text/plain, */*',
    'Origin': 'https://www.shmh.gov.cn/',
    'Referer': 'https://www.shmh.gov.cn/',
}

# --- CẤU HÌNH TỪ KHÓA ---

# BỘ LỌC: Tiêu đề BẮT BUỘC phải chứa ít nhất một trong các từ khóa này
ADDRESS_INDICATOR_KEYWORDS = ["加装电梯工程"]

# BỘ LÀM SẠCH: Các từ thừa ở ĐẦU hoặc CUỐI tiêu đề sẽ bị xóa bỏ
CLEANUP_KEYWORDS = [
    "项目", "地块", "工程", "方案", "规划", "设计", "建设", "加装电梯"
    "新建", "改建", "修缮", "扩建", "改造", "用房", "公示", "公告", "举办", "有关内容予以公示", "工程项目拟报", "工程项目"
    "预公告", "关于", "（已结束）", "的意见征询", "加装电梯", "既有多层住宅", "召开", "专研", "调研", "喜", "好消息", "方案公示"
]


# --- CÁC HÀM HỖ TRỢ ---

def find_last_row_with_data(sheet):
    """Quét ngược từ dưới lên để tìm hàng cuối cùng thực sự có dữ liệu."""
    for row in range(sheet.max_row, 0, -1):
        for cell in sheet[row]:
            if cell.value is not None:
                return row
    return 0


def parse_title_hybrid_improved(title_text):
    """
    Hàm phân tích tiêu đề đã được cải tiến:
    1. Xóa các từ khóa không mong muốn ở đầu chuỗi.
    2. Cắt bỏ phần đuôi không mong muốn bắt đầu bằng một từ khóa.
    3. Tìm từ khóa địa chỉ để cắt chuỗi chính xác.
    """
    address = title_text.strip()

    # --- BƯỚC 1: Xóa các từ khóa thừa ở ĐẦU chuỗi ---
    is_prefix_cleaned = True
    while is_prefix_cleaned:
        is_prefix_cleaned = False

        # FIRST: Remove location prefixes (上海市, 静安区)
        location_prefixes = ["上海市闵行区", "上海闵行区", "闵行区", "上海市", "上海", "闵行"]
        for location in location_prefixes:
            if address.startswith(location):
                address = address[len(location):].strip(' :：')
                is_prefix_cleaned = True
                print(f"  🗑️ Đã xóa địa danh: '{location}'")
                break

        # ONLY if no location was removed, then check CLEANUP_KEYWORDS
        if not is_prefix_cleaned:
            for keyword in CLEANUP_KEYWORDS:
                if address.startswith(keyword):
                    address = address[len(keyword):].strip(' :：')
                    is_prefix_cleaned = True
                    print(f"  🧹 Đã xóa từ khóa: '{keyword}'")
                    break

    # --- BƯỚC 2: Cắt bỏ phần đuôi thừa ---
    min_pos = -1
    for keyword in CLEANUP_KEYWORDS:
        pos = address.find(keyword)
        if pos > 0 and (min_pos == -1 or pos < min_pos):
            min_pos = pos

    if min_pos != -1:
        address = address[:min_pos].strip()

    # --- BƯỚC 3: Tìm và cắt tại từ khóa địa chỉ (FIXED PART) ---
    address = address.replace('（暂名）', '').strip()

    # Search for ALL address keywords, not just "号"
    address_keywords = ["号", "号楼", "单元", "幢）", "楼", "弄", "宅楼", "幢", "街道", "小区", "中学", "小学", "街坊", "村", "苑", "院"]
    found_pos = -1
    found_keyword = None

    for keyword in address_keywords:
        pos = address.rfind(keyword)
        if pos != -1:
            # Calculate where to cut (end of the keyword)
            cut_position = pos + len(keyword)
            # Only use this if it's further right than previous finds
            if cut_position > found_pos:
                found_pos = cut_position
                found_keyword = keyword

    # If we found any address keyword, cut the string there
    if found_pos != -1 and found_keyword:
        address = address[:found_pos].strip()
        print(f"  ✅ Đã cắt tại từ khóa: '{found_keyword}'")

    return "闵行区", address

def read_checkpoint():
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, 'r') as f:
            try:
                content = f.read().strip()
                return int(content) if content else 0
            except (ValueError, IndexError):
                return 0
    return 0

def write_checkpoint(page_num):
    with open(checkpoint_file, 'w') as f:
        f.write(str(page_num))

# --- PHẦN CHÍNH: THU THẬP VÀ XỬ LÝ DỮ LIỆU ---
try:
    all_extracted_data = []
    all_raw_titles = []

    # --- Lấy tổng số trang một cách tự động ---
    print("Đang kiểm tra tổng số trang...")
    initial_payload = {
        "cid": "dOIsK2GVTQfYOJS4EeNHfiaJWNcCku7x", "uid": "dOIsK2GVTQfYOJS4EeNHfiaJWNcCku7x", "query": "加装电梯工程",
        "current": 1, "size": 20, "disable_correction": False,
        "facets": {"view": [{"type": "value", "name": "view", "sort": {"count": "desc"}, "size": 10}]},
        "input_type": "Input"
    }

    try:
        initial_response = requests.post(api_url, headers=headers, json=initial_payload, timeout=30)
        initial_response.raise_for_status()
        initial_data = initial_response.json()
        total_pages = initial_data.get('result', {}).get('_meta', {}).get('page', {}).get('total_pages', 0)
        if total_pages == 0:
            print("Không tìm thấy kết quả hoặc không thể lấy được tổng số trang.")
            exit()
        print(f"Phát hiện thấy có tổng cộng {total_pages} trang kết quả.")
    except requests.exceptions.RequestException as e:
        print(f"Lỗi nghiêm trọng: Không thể lấy thông tin tổng số trang. Lỗi: {e}")
        exit()

    # --- Bắt đầu quá trình crawl ---
    last_completed_page = read_checkpoint()
    start_page = last_completed_page + 1

    if start_page > total_pages:
        print(f"Checkpoint cho thấy đã thu thập xong {last_completed_page}/{total_pages} trang. Không cần chạy lại.")
    else:
        if start_page > 1:
            print(f"Đã phát hiện checkpoint. Tiếp tục từ trang {start_page}...")

        for page_num in range(start_page, total_pages + 1):
            payload = initial_payload.copy()
            payload['current'] = page_num
            print(f"--- Đang thu thập trang {page_num}/{total_pages} từ API ---")

            response = None
            for attempt in range(retry_attempts):
                try:
                    response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                    response.raise_for_status()
                    break
                except requests.exceptions.RequestException as req_err:
                    print(f"  Lỗi kết nối ở trang {page_num} (lần {attempt + 1}/{retry_attempts}): {req_err}")
                    if attempt < retry_attempts - 1:
                        time.sleep(retry_delay)
                    else:
                        print("  Đã hết số lần thử lại. Bỏ qua trang này.")

            if response is None: continue

            try:
                data = response.json()
                results = data.get('result', {}).get('items', [])
            except json.JSONDecodeError:
                print(f"  Không thể phân tích JSON từ trang {page_num}. Bỏ qua."); continue

            if not results: print(f"Trang {page_num} không có dữ liệu. Có thể đã đến trang cuối."); break

            for item in results:
                full_title = item.get('title', {}).get('raw', '').strip()

                # BƯỚC 1: LỌC. Chỉ xử lý nếu tiêu đề chứa từ khóa địa chỉ.
                if not any(keyword in full_title for keyword in ADDRESS_INDICATOR_KEYWORDS):
                    print(f"  -> Bỏ qua (không có từ khóa địa chỉ): {full_title}")
                    continue

                # BƯỚC 2: XỬ LÝ. Nếu đã qua bộ lọc, tiến hành làm sạch và trích xuất.
                date_string = item.get('date', {}).get('raw', '').strip()

                if full_title and date_string:
                    all_raw_titles.append(full_title)
                    district, address = parse_title_hybrid_improved(full_title) # Gọi hàm làm sạch nâng cao
                    parts = date_string.split('-')
                    year, month, day = (parts[0], parts[1], parts[2]) if len(parts) == 3 else ("", "", "")
                    all_extracted_data.append({
                        '区': district, '地址': address, '年': year, '月': month, '日': day
                    })

            write_checkpoint(page_num)
            print(f"  Đã xử lý xong trang {page_num}. Đã lưu checkpoint.")
            time.sleep(request_delay)

        # --- Ghi dữ liệu ra file ---
        print(f"\n>>> Thu thập hoàn tất. Tổng cộng {len(all_extracted_data)} tiêu đề hợp lệ đã được xử lý trong lần chạy này. <<<\n")

        if all_raw_titles:
            with open(txt_filename, 'a', encoding='utf-8') as f:
                f.write('\n'.join(all_raw_titles) + '\n')
            print(f"Đã thêm {len(all_raw_titles)} tiêu đề thô vào file '{txt_filename}'")

        if all_extracted_data:
            columns_order = ['区', '地址', '年', '月', '日']
            new_df = pd.DataFrame(all_extracted_data)[columns_order]
            try:
                book = load_workbook(excel_filename)
                sheet = book[sheet_name_to_update] if sheet_name_to_update in book.sheetnames else book.create_sheet(sheet_name_to_update)
                last_data_row = find_last_row_with_data(sheet)
                start_row = last_data_row + 1

                # Ghi header nếu là sheet mới
                if start_row <= 1:
                    headers_list = list(new_df.columns)
                    for col_idx, header_value in enumerate(headers_list, start=2): sheet.cell(row=1, column=col_idx, value=header_value)
                    start_row = 2 # Dữ liệu sẽ bắt đầu từ hàng 2

                print(f"Đang ghi {len(new_df)} dòng dữ liệu mới vào sheet '{sheet_name_to_update}' bắt đầu từ hàng {start_row}...")
                for i, row_data in new_df.iterrows():
                    current_row = start_row + i
                    for col_idx, col_name in enumerate(columns_order, start=2):
                        sheet.cell(row=current_row, column=col_idx, value=row_data[col_name])

                book.save(excel_filename)
                print("Hoàn tất! File Excel đã được cập nhật thành công.")
            except FileNotFoundError: print(f"Lỗi: Không tìm thấy file Excel '{excel_filename}'.")
            except Exception as ex: print(f"Lỗi khi ghi file Excel: {ex}")
        else:
            print("Không có dữ liệu mới hợp lệ để ghi vào file Excel.")
except Exception as e:
    print(f"Đã xảy ra một lỗi không mong muốn: {e}")