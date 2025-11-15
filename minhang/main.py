import requests
import pandas as pd
import os
import re
import time
from openpyxl import load_workbook
import json

# --- CẤU HÌNH BỘ LỌC THÔNG MINH ---
# (Các danh sách từ khóa này vẫn rất hữu ích và được giữ nguyên)
ADDRESS_INDICATOR_KEYWORDS = ["加装电梯工程"]
CLEANUP_KEYWORDS = [
    "项目", "地块", "工程", "方案", "规划", "设计", "建设", "加装电梯",
    "新建", "改建", "修缮", "扩建", "改造", "用房", "公示", "公告", "举办",
    "有关内容予以公示", "工程项目拟报", "工程项目", "预公告", "关于", "（已结束）",
    "的意见征询", "既有多层住宅", "召开", "专研", "调研", "喜", "好消息", "方案公示"
]

# --- CẤU HÌNH KỸ THUẬT ---
# <<< THAY ĐỔI 1: Cập nhật API URL và các tham số cơ bản >>>
base_api_url = r'https://zwgk.shmh.gov.cn/mh-xxgk-cms/searchindex/searchByDB'
# Các tham số cố định
base_params = {
    'keyword': '加装电梯工程',
    'sitecode': 'mh_xxgk',
    'sitename': '闵行信息公开',
    'pagesize': 10
}

request_delay = 0.5
retry_attempts = 3
retry_delay = 5

checkpoint_file = 'checkpoint.log'
txt_filename = 'crawled_titles.txt'
excel_filename = '模版.xlsx'
sheet_name_to_update = 'shanghai'

# Headers cho trang này có thể đơn giản hơn
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
}

# --- CÁC HÀM HỖ TRỢ ---
# (Các hàm này đã ổn định, chỉ sửa lại tên quận trong parse_title)
def find_last_row_with_data(sheet):
    for row in range(sheet.max_row, 0, -1):
        for cell in sheet[row]:
            if cell.value is not None: return row
    return 0

def parse_title_hybrid_improved(title_text):
    address = title_text.strip()
    location_prefixes = ["上海市闵行区", "上海闵行区", "闵行区", "上海市", "上海", "闵行"]
    for location in location_prefixes:
        if address.startswith(location):
            address = address[len(location):].strip(' :：')
            break
    min_pos = -1
    for keyword in CLEANUP_KEYWORDS:
        pos = address.find(keyword)
        if pos > 0 and (min_pos == -1 or pos < min_pos): min_pos = pos
    if min_pos != -1: address = address[:min_pos].strip()
    address_keywords = ["号", "号楼", "单元", "幢）", "楼", "弄", "宅楼", "幢", "街道", "小区", "中学", "小学", "街坊", "村", "苑", "院"]
    found_pos = -1
    for keyword in address_keywords:
        pos = address.rfind(keyword)
        if pos != -1 and (pos + len(keyword)) > found_pos:
            found_pos = pos + len(keyword)
    if found_pos != -1: address = address[:found_pos].strip()
    return "闵行区", address

def read_checkpoint():
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, 'r') as f:
            try: return int(f.read().strip())
            except ValueError: return 0
    return 0

def write_checkpoint(page_num):
    with open(checkpoint_file, 'w') as f: f.write(str(page_num))

# --- PHẦN CHÍNH: THU THẬP VÀ XỬ LÝ DỮ LIỆU ---
try:
    all_extracted_data = []; all_raw_titles = []

    print("Đang kiểm tra tổng số trang...")
    
    # <<< THAY ĐỔI 2: Cách lấy tổng số trang cho API GET >>>
    try:
        # Gửi request cho trang đầu tiên để lấy tổng số kết quả
        params = base_params.copy()
        params['pageindex'] = 1
        params['t'] = int(time.time() * 1000) # Tạo timestamp mới
        
        initial_response = requests.get(base_api_url, headers=headers, params=params, timeout=30)
        initial_response.raise_for_status()
        initial_data = initial_response.json()
        
        total_count = initial_data.get('totalcount', 0)
        page_size = initial_data.get('pagesize', 10)
        if total_count == 0:
            print("Không tìm thấy kết quả hoặc không thể lấy được tổng số trang.")
            exit()
        
        # Tính toán tổng số trang
        total_pages = (total_count + page_size - 1) // page_size
        print(f"Phát hiện thấy có tổng cộng {total_count} kết quả ({total_pages} trang).")
    except requests.exceptions.RequestException as e:
        print(f"Lỗi nghiêm trọng: Không thể lấy thông tin tổng số trang. Lỗi: {e}")
        exit()

    last_completed_page = read_checkpoint()
    start_page = last_completed_page + 1

    if start_page > total_pages:
        print(f"Checkpoint cho thấy đã thu thập xong {last_completed_page}/{total_pages} trang.")
    else:
        if start_page > 1:
            print(f"Đã phát hiện checkpoint. Tiếp tục từ trang {start_page}...")

        # <<< THAY ĐỔI 3: Vòng lặp chính sử dụng API GET >>>
        for page_num in range(start_page, total_pages + 1):
            
            # Tạo các tham số cho trang hiện tại
            current_params = base_params.copy()
            current_params['pageindex'] = page_num
            current_params['t'] = int(time.time() * 1000) # Luôn tạo timestamp mới cho mỗi request

            print(f"--- Đang thu thập trang {page_num}/{total_pages} từ API ---")

            response = None
            for attempt in range(retry_attempts):
                try:
                    response = requests.get(base_api_url, headers=headers, params=current_params, timeout=30)
                    response.raise_for_status()
                    break
                except requests.exceptions.RequestException as req_err:
                    print(f"  Lỗi kết nối (lần {attempt + 1}): {req_err}"); time.sleep(retry_delay)
            if response is None: continue

            try:
                data = response.json()
                # Lấy danh sách kết quả từ khóa 'list'
                results = data.get('list', [])
            except json.JSONDecodeError: print(f"  Lỗi JSON. Bỏ qua."); continue

            if not results: print(f"Trang {page_num} không có dữ liệu."); break

            # <<< THAY ĐỔI 4: Trích xuất dữ liệu từ cấu trúc JSON mới >>>
            for item in results:
                # Dọn dẹp HTML tag khỏi tiêu đề
                full_title_html = item.get('title', '').strip()
                full_title = re.sub('<[^>]+>', '', full_title_html) # Xóa các thẻ HTML như <span>

                if not any(keyword in full_title for keyword in ADDRESS_INDICATOR_KEYWORDS):
                    print(f"  -> Bỏ qua: {full_title}")
                    continue

                date_string = item.get('startTime', '').strip()
                
                if full_title and date_string:
                    all_raw_titles.append(full_title)
                    district, address = parse_title_hybrid_improved(full_title)
                    parts = date_string.split('-')
                    year, month, day = (parts[0], parts[1], parts[2]) if len(parts) == 3 else ("", "", "")
                    all_extracted_data.append({'区': district, '地址': address, '年': year, '月': month, '日': day})

            write_checkpoint(page_num)
            print(f"  Đã xử lý xong trang {page_num}. Đã lưu checkpoint.")
            time.sleep(request_delay)
            
        # (Phần ghi dữ liệu không thay đổi)
        print(f"\n>>> HOÀN TẤT. {len(all_extracted_data)} mục hợp lệ đã được xử lý. <<<\n")
        if all_extracted_data:
            columns_order = ['区', '地址', '年', '月', '日']
            new_df = pd.DataFrame(all_extracted_data)[columns_order]
            try:
                book = load_workbook(excel_filename)
                sheet = book[sheet_name_to_update] if sheet_name_to_update in book.sheetnames else book.create_sheet(sheet_name_to_update)
                last_data_row = find_last_row_with_data(sheet)
                start_row = last_data_row + 1
                if start_row <= 1:
                    for col_idx, val in enumerate(list(new_df.columns), start=2): sheet.cell(row=1, column=col_idx, value=val)
                    start_row = 2
                print(f"Đang ghi dữ liệu vào Excel...")
                for i, row_data in new_df.iterrows():
                    for col_idx, col_name in enumerate(columns_order, start=2):
                        sheet.cell(row=start_row + i, column=col_idx, value=row_data[col_name])
                book.save(excel_filename)
                print("Đã lưu file Excel.")
            except FileNotFoundError: print(f"Lỗi: Không tìm thấy file Excel '{excel_filename}'.")
            except Exception as ex: print(f"Lỗi Excel: {ex}")
except Exception as e:
    print(f"Đã xảy ra một lỗi không mong muốn: {e}")