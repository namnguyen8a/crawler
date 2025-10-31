import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import re
import time
from openpyxl import load_workbook

# --- CẤU HÌNH ---
base_url = 'https://www.shyp.gov.cn/shypq/yqyw-wb-gtjzl-gsgg-fags/'
total_pages = 251
request_delay = 0.5 

checkpoint_file = 'checkpoint.log'
retry_attempts = 3
retry_delay = 5

txt_filename = 'crawled_titles.txt'
excel_filename = '模版.xlsx'
sheet_name_to_update = 'shanghai'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

# (Các hàm tiện ích giữ nguyên)
def find_last_row_with_data(sheet):
    for row in range(sheet.max_row, 0, -1):
        for cell in sheet[row]:
            if cell.value is not None:
                return row
    return 0

def parse_title_hybrid(title_text):
    district = ""
    address = title_text.strip()
    match = re.search(r'(杨浦区)', address)
    if match:
        district = match.group(1)
        address = address.replace(district, '', 1).strip()
    address = address.replace('（暂名）', '').strip()
    BOUNDARY_KEYWORDS = [
        "项目", "地块", "工程", "方案", "规划", "设计", "建设", "新建", "改建", "修缮", "扩建", "改造", "用房", "公示", "公告"
    ]
    min_pos = -1
    for keyword in BOUNDARY_KEYWORDS:
        pos = address.find(keyword)
        if pos != -1 and (min_pos == -1 or pos < min_pos):
            min_pos = pos
    if min_pos != -1:
        address = address[:min_pos].strip()
    return district, address

def read_checkpoint():
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, 'r') as f:
            try: return int(f.read().strip())
            except ValueError: return 0
    return 0

def write_checkpoint(page_num):
    with open(checkpoint_file, 'w') as f:
        f.write(str(page_num))

# --- BẮT ĐẦU SCRIPT ---
try:
    all_extracted_data = []
    all_raw_titles = []

    # Đọc checkpoint để xác định trang bắt đầu
    last_completed_page = read_checkpoint()
    start_page = last_completed_page + 1

    # <<< THAY ĐỔI QUAN TRỌNG: Kiểm tra xem đã hoàn thành chưa >>>
    if start_page > total_pages:
        print(f"Checkpoint cho thấy đã thu thập xong {last_completed_page}/{total_pages} trang. Không cần chạy lại.")
        print("Để thu thập lại từ đầu, vui lòng xóa file 'checkpoint.log' và chạy lại script.")
    else:
        # Chỉ chạy vòng lặp nếu công việc chưa hoàn thành
        if start_page > 1:
            print(f"Đã phát hiện checkpoint. Tiếp tục từ trang {start_page}...")

        for page_num in range(start_page, total_pages + 1):
            if page_num == 1:
                current_url = base_url + "index.html"
            else:
                current_url = f"{base_url}index_{page_num}.html"
            
            print(f"--- Đang thu thập trang {page_num}/{total_pages}: {current_url} ---")
            
            response = None
            for attempt in range(retry_attempts):
                try:
                    response = requests.get(current_url, headers=headers, timeout=30)
                    response.raise_for_status()
                    break
                except requests.exceptions.HTTPError as http_err:
                    print(f"  Lỗi HTTP khi truy cập trang {page_num}: {http_err}. Có thể trang không tồn tại. Dừng lại.")
                    response = None
                    break
                except requests.exceptions.RequestException as req_err:
                    print(f"  Lỗi kết nối ở trang {page_num} (lần {attempt + 1}/{retry_attempts}): {req_err}")
                    if attempt < retry_attempts - 1:
                        print(f"  Sẽ thử lại sau {retry_delay} giây...")
                        time.sleep(retry_delay)
                    else:
                        print("  Đã hết số lần thử lại. Bỏ qua trang này.")
            
            if response is None:
                continue
            
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            list_items = soup.select('ul.uli16.nowrapli.padding-top-10.list-date.border li')

            if not list_items:
                print(f"Trang {page_num} không có dữ liệu. Dừng quá trình thu thập.")
                break

            for item in list_items:
                date_tag = item.find('span', class_='time')
                title_tag = item.find('a')
                if date_tag and title_tag:
                    full_title = title_tag.get_text(strip=True)
                    all_raw_titles.append(full_title)
                    district, address = parse_title_hybrid(full_title)
                    date_string = date_tag.get_text(strip=True)
                    parts = date_string.split('.')
                    year, month, day = (parts[0], parts[1], parts[2]) if len(parts) == 3 else ("", "", "")
                    all_extracted_data.append({
                        '区': district, '地址': address, '年': year, '月': month, '日': day
                    })
            
            write_checkpoint(page_num)
            print(f"  Đã xử lý xong trang {page_num}. Đã lưu checkpoint.")
            time.sleep(request_delay)

        # --- KẾT THÚC VÒNG LẶP ---

        print(f"\n>>> Thu thập hoàn tất. Tổng cộng {len(all_raw_titles)} tiêu đề mới đã được lấy trong lần chạy này. <<<\n")

        # Ghi file txt ở chế độ 'a' (append)
        if all_raw_titles:
            with open(txt_filename, 'a', encoding='utf-8') as f:
                f.write('\n'.join(all_raw_titles) + '\n')
            print(f"Hoàn tất! Đã thêm {len(all_raw_titles)} tiêu đề thô vào file '{txt_filename}'")
        
        # Chỉ ghi vào Excel nếu có dữ liệu mới
        if all_extracted_data:
            columns_order = ['区', '地址', '年', '月', '日']
            new_df = pd.DataFrame(all_extracted_data)[columns_order]
            book = load_workbook(excel_filename)
            if sheet_name_to_update in book.sheetnames:
                sheet = book[sheet_name_to_update]
            else:
                sheet = book.create_sheet(sheet_name_to_update)
                # Chỉ ghi header nếu sheet thực sự trống
                if find_last_row_with_data(sheet) == 0:
                    headers_list = list(new_df.columns)
                    for col_idx, header_value in enumerate(headers_list):
                        sheet.cell(row=1, column=col_idx + 2, value=header_value)
            
            last_data_row = find_last_row_with_data(sheet)
            start_row = last_data_row + 1

            print(f"Đang ghi {len(new_df)} dòng dữ liệu mới vào sheet '{sheet_name_to_update}' bắt đầu từ hàng {start_row}...")
            for i, row_data in new_df.iterrows():
                current_row = start_row + i
                sheet.cell(row=current_row, column=2, value=row_data['区'])
                sheet.cell(row=current_row, column=3, value=row_data['地址'])
                sheet.cell(row=current_row, column=4, value=row_data['年'])
                sheet.cell(row=current_row, column=5, value=row_data['月'])
                sheet.cell(row=current_row, column=6, value=row_data['日'])
            
            book.save(excel_filename)
            print("Hoàn tất! File đã được cập nhật thành công.")
        else:
            print("Không có dữ liệu mới để ghi vào file Excel.")

except Exception as e:
    print(f"Đã xảy ra một lỗi không mong muốn: {e}")