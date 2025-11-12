import requests
import pandas as pd
import os
import re
import time
from openpyxl import load_workbook
import json

# --- CẤU HÌNH CHÍNH ---

# 1. TỪ KHÓA TÌM KIẾM (Query gửi lên API)
SEARCH_KEYWORDS = ["加装电梯"] 

# 2. BỘ LỌC "QUY TẮC VÀNG" (Regex Whitelist)
# Tiêu đề PHẢI khớp với ít nhất một trong các mẫu này mới được coi là hợp lệ.
# \d+ nghĩa là "một hoặc nhiều chữ số".
ADDRESS_REGEX_PATTERN = re.compile(
    r'\d+号楼?'   # Ví dụ: 15号 hoặc 15号楼
    r'|\d+弄'     # Ví dụ: 123弄
    r'|\d+幢'     # Ví dụ: 5幢
    r'|\d+支弄'   # Ví dụ: 1支弄
    r'|\d+街坊'   # Ví dụ: 96街坊
)

# 3. TỪ KHÓA LOẠI TRỪ (Blacklist) - Dùng để loại bỏ tin tức, rác
BLACKLIST_KEYWORDS = [
    "培训", "会议", "宣传", "大民生", "指导意见", "一图读懂", "工作推进", "约法三章", 
    "指尖办理", "后半篇文章", "提质增效", "印发", "通知", "民心工程", "结合", 
    "治理", "党建", "实事", "全覆盖", "如何", "推进会", "经验", "赛跑", "负责"
]

# 4. TỪ KHÓA LÀM SẠCH (Cleanup) - Dùng để cắt bỏ phần thừa SAU KHI đã lọc
CLEANUP_KEYWORDS = [
    "街道", "项目", "地块", "工程", "方案", "规划", "设计", "建设", "新建", "改建", "修缮", 
    "扩建", "改造", "用房", "公示", "公告", "预公告", "关于", "（已结束）", "的意见征询", 
    "旧住房", "成套", "公众反馈意见的处理情况"
]

# --- CẤU HÌNH KỸ THUẬT ---
# <<< KHÔI PHỤC LẠI THÔNG SỐ CHÍNH XÁC CỦA BẠN >>>
api_url = r'https://ss.shanghai.gov.cn/manda-app/api/app/search/v1/1ywaiqo/search'
request_delay = 0.5; retry_attempts = 3; retry_delay = 5
checkpoint_file = 'checkpoint.log'; txt_filename = 'crawled_titles.txt'
excel_filename = '模版.xlsx'; sheet_name_to_update = 'shanghai'

# Giữ nguyên headers của bạn
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36',
    'Content-Type': 'application/json',
    'Accept': 'application/json, text/plain, */*',
    'Origin': 'https://www.shhuangpu.gov.cn',
    'Referer': 'https://www.shhuangpu.gov.cn/',
}

# --- CÁC HÀM HỖ TRỢ ---
# (Các hàm này đã ổn định, không cần thay đổi)
def find_last_row_with_data(sheet):
    for row in range(sheet.max_row, 0, -1):
        for cell in sheet[row]:
            if cell.value is not None: return row
    return 0
def parse_title_hybrid_smart(title_text):
    address = title_text.strip()
    is_prefix_cleaned = True
    while is_prefix_cleaned:
        is_prefix_cleaned = False
        for keyword in CLEANUP_KEYWORDS:
            if address.startswith(keyword):
                address = address[len(keyword):].strip(' :：'); is_prefix_cleaned = True; break
    address = address.replace('（暂名）', '').strip()
    hao_lou_pos = address.rfind('号楼')
    if hao_lou_pos != -1: return "黄浦区", address[:hao_lou_pos + 2].strip()
    hao_pos = address.rfind('号')
    if hao_pos != -1: return "黄浦区", address[:hao_pos + 1].strip()
    min_pos = -1
    for keyword in CLEANUP_KEYWORDS:
        pos = address.find(keyword)
        if pos > 0 and (min_pos == -1 or pos < min_pos): min_pos = pos
    if min_pos != -1: address = address[:min_pos].strip()
    return "黄浦区", address
def read_checkpoint():
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, 'r', encoding='utf-8') as f:
            try:
                content = f.read().strip().split(',');
                if len(content) == 2: return content[0], int(content[1])
            except (ValueError, IndexError): return None, 0
    return None, 0
def write_checkpoint(keyword, page_num):
    with open(checkpoint_file, 'w', encoding='utf-8') as f: f.write(f"{keyword},{page_num}")

# --- PHẦN CHÍNH ---
try:
    all_extracted_data = []; all_raw_titles = []
    saved_keyword, last_completed_page = read_checkpoint()

    for keyword in SEARCH_KEYWORDS:
        if saved_keyword and SEARCH_KEYWORDS.index(keyword) < SEARCH_KEYWORDS.index(saved_keyword):
            print(f"Bỏ qua: '{keyword}'"); continue

        print(f"\n{'='*20} BẮT ĐẦU VỚI TỪ KHÓA: '{keyword}' {'='*20}")
        
        # <<< KHÔI PHỤC LẠI PAYLOAD CHÍNH XÁC CỦA BẠN >>>
        initial_payload = {
            "cid": "vcXSblaG6SxswWniwgOsNTFg1qcaNlvo",
            "uid": "vcXSblaG6SxswWniwgOsNTFg1qcaNlvo",
            "query": keyword,
            "current": 1,
            "size": 20, # Kích thước trang là 20
            "disable_correction": False,
            "facets": {"view": [{"type": "value", "name": "view", "sort": {"count": "desc"}, "size": 10}]},
            "input_type": "Input"
        }
        
        try:
            print("Đang kiểm tra tổng số trang...")
            initial_response = requests.post(api_url, headers=headers, json=initial_payload, timeout=30)
            initial_response.raise_for_status()
            initial_data = initial_response.json()
            total_pages = initial_data.get('result', {}).get('_meta', {}).get('page', {}).get('total_pages', 0)
            if total_pages == 0: print(f"Không có kết quả. Bỏ qua."); continue
            print(f"Phát hiện {total_pages} trang kết quả.")
        except requests.exceptions.RequestException as e: print(f"Lỗi khi lấy tổng số trang: {e}"); continue
        
        start_page = 1
        if keyword == saved_keyword: start_page = last_completed_page + 1
        if start_page > total_pages: print(f"Đã xong từ khóa '{keyword}'."); continue

        for page_num in range(start_page, total_pages + 1):
            payload = initial_payload.copy(); payload['current'] = page_num
            print(f"--- Đang thu thập trang {page_num}/{total_pages} (Từ khóa: '{keyword}') ---")
            
            response = None
            for attempt in range(retry_attempts):
                try:
                    response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                    response.raise_for_status(); break
                except requests.exceptions.RequestException as req_err:
                    print(f"  Lỗi kết nối (lần {attempt + 1}): {req_err}"); time.sleep(retry_delay)
            if response is None: continue
            
            try: data = response.json(); results = data.get('result', {}).get('items', [])
            except json.JSONDecodeError: print(f"  Lỗi JSON."); continue
            if not results: print(f"Trang {page_num} không có dữ liệu."); break

            for item in results:
                full_title = item.get('title', {}).get('raw', '').strip()
                if not full_title: continue

                # ============================================================
                # BỘ LỌC REGEX THÔNG MINH
                # ============================================================
                # 1. KIỂM TRA BLACKLIST
                if any(spam_word in full_title for spam_word in BLACKLIST_KEYWORDS):
                    print(f"  [BỎ QUA - TIN TỨC] {full_title}"); continue

                # 2. KIỂM TRA "QUY TẮC VÀNG"
                if not ADDRESS_REGEX_PATTERN.search(full_title):
                    print(f"  [BỎ QUA - KHÔNG CÓ MẪU ĐỊA CHỈ] {full_title}"); continue
                
                # NẾU VƯỢT QUA CẢ 2 BỘ LỌC -> CHẤP NHẬN
                print(f"  [CHẤP NHẬN] {full_title}")
                all_raw_titles.append(full_title)
                district, address = parse_title_hybrid_smart(full_title)
                date_string = item.get('date', {}).get('raw', '').strip()
                parts = date_string.split('-')
                year, month, day = (parts[0], parts[1], parts[2]) if len(parts) == 3 else ("", "", "")
                all_extracted_data.append({'区': district, '地址': address, '年': year, '月': month, '日': day})

            write_checkpoint(keyword, page_num)
            time.sleep(request_delay)
    
    # --- GHI DỮ LIỆU ---
    # (Phần này không thay đổi)
    print(f"\n>>> HOÀN TẤT. {len(all_extracted_data)} mục hợp lệ đã được xử lý. <<<\n")
    if all_extracted_data:
        columns_order = ['区', '地址', '年', '月', '日']
        new_df = pd.DataFrame(all_extracted_data)[columns_order]
        try:
            book = load_workbook(excel_filename)
            sheet = book[sheet_name_to_update] if sheet_name_to_update in book.sheetnames else book.create_sheet(sheet_name_to_update)
            last_data_row = find_last_row_with_data(sheet)
            start_row = last_data_row + 1
            if start_row <= 1:
                for col_idx, val in enumerate(list(new_df.columns), start=2): sheet.cell(row=1, column=col_idx, value=val)
                start_row = 2
            print(f"Đang ghi dữ liệu vào Excel...")
            for i, row_data in new_df.iterrows():
                for col_idx, col_name in enumerate(columns_order, start=2):
                    sheet.cell(row=start_row + i, column=col_idx, value=row_data[col_name])
            book.save(excel_filename)
            print("Đã lưu file Excel.")
        except FileNotFoundError: print(f"Lỗi: Không tìm thấy file Excel '{excel_filename}'.")
        except Exception as ex: print(f"Lỗi Excel: {ex}")

except Exception as e:
    print(f"Đã xảy ra lỗi: {e}")