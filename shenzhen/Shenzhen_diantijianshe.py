import requests
import pandas as pd
import os
import re
import time
from openpyxl import load_workbook
import json

# --- CẤU HÌNH ---
# API endpoint chính xác
api_url = 'https://search.gd.gov.cn/api/search/all'

# Các biến điều khiển quá trình crawl
request_delay = 0.5
retry_attempts = 3
retry_delay = 5

# Tên các file output và checkpoint
checkpoint_file = 'checkpoint_shenzhen.log'
txt_filename = 'crawled_titles_shenzhen.txt'
excel_filename = '251107_SH Crawled Data File.xlsx'
sheet_name_to_update = 'shenzhen'

# Headers
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Content-Type': 'application/json',
    'Accept': 'application/json, text/plain, */*',
    'Origin': 'https://search.gd.gov.cn',
    'Referer': 'https://search.gd.gov.cn/',
}

# --- CẤU HÌNH TỪ KHÓA ---
ADDRESS_INDICATOR_KEYWORDS = ["电梯建设", "电梯工程", "电梯总平"]
CLEANUP_KEYWORDS = [
    "项目", "地块", "工程", "方案", "规划", "设计", "建设", "深圳市", "南山区", "设计方案", "总平面",
    "新建", "改建", "修缮", "扩建", "改造", "用房", "公示", "公告", "关于", "通告", "加建电梯", "新增",
    "批前公示", "规划许可", "管理局", "市规划和自然资源局", "南山管理局", "新增电梯工程总平面图修改的公示", "加建电梯总平面图的通告"
]

# Danh sách các quận tại Thâm Quyến
SHENZHEN_DISTRICTS = [
    "南山区", "福田区", "罗湖区", "盐田区", "宝安区", "龙岗区", "龙华区", "坪山区", "光明区", "大鹏新区"
]

# Mapping từ khóa địa danh đến quận
DISTRICT_KEYWORDS = {
    "南山区": ["南山", "蛇口", "华侨城", "科技园", "后海", "前海", "西丽", "粤海", "沙河"],
    "福田区": ["福田", "华强北", "中心区", "皇岗", "莲花山", "香蜜湖", "梅林", "园岭"],
    "罗湖区": ["罗湖", "东门", "国贸", "火车站", "莲塘", "黄贝", "桂园", "笋岗"],
    "盐田区": ["盐田", "大梅沙", "小梅沙", "沙头角", "海山"],
    "宝安区": ["宝安", "西乡", "福永", "沙井", "松岗", "石岩", "新安", "航城"],
    "龙岗区": ["龙岗", "布吉", "横岗", "平湖", "坂田", "南湾", "园山", "宝龙"],
    "龙华区": ["龙华", "观澜", "民治", "大浪", "福城", "观湖"],
    "坪山区": ["坪山", "坑梓", "龙田", "石井", "马峦"],
    "光明区": ["光明", "公明", "新湖", "凤凰", "玉塘", "马田"],
    "大鹏新区": ["大鹏", "葵涌", "南澳", "大鹏街道", "葵涌街道"]
}


# --- CÁC HÀM HỖ TRỢ ---
def find_last_row_with_data(sheet):
    """Quét ngược từ dưới lên để tìm hàng cuối cùng thực sự có dữ liệu."""
    for row in range(sheet.max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row, col).value is not None:
                return row
    return 0


def extract_district_from_title(title):
    """
    Trích xuất quận trực tiếp từ tiêu đề (ưu tiên cao nhất)
    """
    # Tìm quận trực tiếp trong tiêu đề
    for district in SHENZHEN_DISTRICTS:
        if district in title:
            return district

    # Tìm từ khóa quận trong tiêu đề (ví dụ: "福田管理局" -> "福田区")
    for district, keywords in DISTRICT_KEYWORDS.items():
        for keyword in keywords:
            # Tìm pattern như "福田管理局", "南山管理局", etc.
            if f"{keyword}管理局" in title:
                return district
            # Tìm từ khóa đơn lẻ
            if keyword in title and len(keyword) > 1:  # Tránh trùng lặp với từ ngắn
                # Kiểm tra xem từ khóa có phải là một từ độc lập không
                pattern = r'[^a-zA-Z0-9]' + re.escape(keyword) + r'[^a-zA-Z0-9]'
                if re.search(pattern, title):
                    return district

    return None


def extract_district_from_content(content, title):
    """
    Trích xuất quận từ nội dung nếu không tìm thấy trong tiêu đề
    """
    # Kết hợp nội dung và tiêu đề để tìm quận
    combined_text = title + " " + content

    # Tìm các quận trong văn bản
    for district in SHENZHEN_DISTRICTS:
        if district in combined_text:
            return district

    # Tìm theo từ khóa địa danh
    for district, keywords in DISTRICT_KEYWORDS.items():
        for keyword in keywords:
            if keyword in combined_text:
                # Kiểm tra xem từ khóa có phải là một từ độc lập không
                pattern = r'[^a-zA-Z0-9]' + re.escape(keyword) + r'[^a-zA-Z0-9]'
                if re.search(pattern, combined_text):
                    return district

    # Nếu vẫn không tìm thấy, trả về "未知区" (Quận không xác định)
    return "未知区"


def get_district(title, content):
    """
    Lấy quận: ưu tiên tiêu đề trước, sau đó đến nội dung
    """
    # Ưu tiên 1: Tìm trong tiêu đề
    district_from_title = extract_district_from_title(title)
    if district_from_title:
        print(f"  🎯 Quận tìm thấy trong tiêu đề: {district_from_title}")
        return district_from_title

    # Ưu tiên 2: Tìm trong nội dung
    district_from_content = extract_district_from_content(content, title)
    if district_from_content and district_from_content != "未知区":
        print(f"  📄 Quận tìm thấy trong nội dung: {district_from_content}")
        return district_from_content

    # Không tìm thấy
    print("  ⚠️  Không thể xác định quận")
    return "未知区"


def parse_address_components(address_text):
    """
    Phân tích địa chỉ thành các thành phần: 小区, 栋数, 单元
    """
    # Làm sạch địa chỉ trước
    address = address_text.strip()

    # --- BƯỚC 1: Xóa các từ khóa thừa ở ĐẦU chuỗi ---
    is_prefix_cleaned = True
    while is_prefix_cleaned:
        is_prefix_cleaned = False
        location_prefixes = ["深圳市", "深圳"]
        for location in location_prefixes:
            if address.startswith(location):
                address = address[len(location):].strip(' :：')
                is_prefix_cleaned = True
                break

        if not is_prefix_cleaned:
            for keyword in CLEANUP_KEYWORDS:
                if address.startswith(keyword):
                    address = address[len(keyword):].strip(' :：')
                    is_prefix_cleaned = True
                    break

    # --- BƯỚC 2: Tách thành phần địa chỉ ---
    xiaoqu = ""  # 小区
    dongshu = ""  # 栋数
    danyuan = ""  # 单元

    # Mẫu regex để tách địa chỉ
    patterns = [
        # Mẫu: "小区名" + "栋/号楼" + "单元"
        r'(.+?)(\d+[栋号楼])(\d*单元)?',
        # Mẫu: "小区名" + "栋"
        r'(.+?)(\d+[栋号楼])',
        # Mẫu: "小区名" + "单元"
        r'(.+?)(\d*单元)',
        # Mẫu chỉ có số (cho các trường hợp đơn giản)
        r'(\D+?)(\d+)'
    ]

    for pattern in patterns:
        match = re.search(pattern, address)
        if match:
            groups = match.groups()
            if len(groups) >= 2:
                xiaoqu = groups[0].strip()
                if '栋' in str(groups[1]) or '号楼' in str(groups[1]) or '楼' in str(groups[1]):
                    dongshu = groups[1].strip()
                else:
                    # Nếu không phải số tòa nhà, có thể là số đơn vị
                    danyuan = groups[1].strip()

                if len(groups) >= 3 and groups[2]:
                    danyuan = groups[2].strip()
                break

    # Nếu không tách được bằng regex, thử phương pháp đơn giản hơn
    if not xiaoqu:
        # Tìm vị trí của các từ khóa địa chỉ
        building_keywords = ["栋", "号楼", "楼", "幢"]
        unit_keywords = ["单元", "座"]

        building_pos = -1
        unit_pos = -1

        for keyword in building_keywords:
            pos = address.find(keyword)
            if pos != -1 and (building_pos == -1 or pos < building_pos):
                building_pos = pos

        for keyword in unit_keywords:
            pos = address.find(keyword)
            if pos != -1 and (unit_pos == -1 or pos < unit_pos):
                unit_pos = pos

        if building_pos != -1:
            xiaoqu = address[:building_pos].strip()
            if unit_pos != -1 and unit_pos > building_pos:
                dongshu = address[building_pos:unit_pos + 2].strip()  # +2 để lấy cả từ khóa
                danyuan = address[unit_pos:].strip()
            else:
                dongshu = address[building_pos:].strip()
        else:
            # Nếu không tìm thấy từ khóa, coi toàn bộ là tên khu phố
            xiaoqu = address

    # Làm sạch kết quả
    xiaoqu = xiaoqu.strip(' ,，.。')
    dongshu = dongshu.strip(' ,，.。')
    danyuan = danyuan.strip(' ,，.。')

    print(f"  🏘️ 小区: '{xiaoqu}'")
    print(f"  🏢 栋数: '{dongshu}'")
    print(f"  🚪 单元: '{danyuan}'")

    return xiaoqu, dongshu, danyuan


def parse_title_hybrid_improved(title_text):
    """
    Hàm phân tích tiêu đề đã được cải tiến cho website Guangdong
    """
    address = title_text.strip()

    # --- BƯỚC 1: Xóa các từ khóa thừa ở ĐẦU chuỗi ---
    is_prefix_cleaned = True
    while is_prefix_cleaned:
        is_prefix_cleaned = False
        location_prefixes = ["深圳市", "深圳"]
        for location in location_prefixes:
            if address.startswith(location):
                address = address[len(location):].strip(' :：')
                is_prefix_cleaned = True
                print(f"  🗑️ Đã xóa địa danh: '{location}'")
                break

        if not is_prefix_cleaned:
            for keyword in CLEANUP_KEYWORDS:
                if address.startswith(keyword):
                    address = address[len(keyword):].strip(' :：')
                    is_prefix_cleaned = True
                    print(f"  🧹 Đã xóa từ khóa: '{keyword}'")
                    break

    # --- BƯỚC 2: Cắt bỏ phần đuôi thừa ---
    min_pos = -1
    for keyword in CLEANUP_KEYWORDS:
        pos = address.find(keyword)
        if pos > 0 and (min_pos == -1 or pos < min_pos):
            min_pos = pos

    if min_pos != -1:
        address = address[:min_pos].strip()

    # --- BƯỚC 3: Tìm và cắt tại từ khóa địa chỉ ---
    address = address.replace('（暂名）', '').strip()
    address_keywords = ["号", "单元", "号楼", "楼", "弄", "宅楼", "幢", "街道", "小区", "中学", "小学", "街坊", "村",
                        "苑", "院", "栋", "花园"]
    found_pos = -1
    found_keyword = None

    for keyword in address_keywords:
        pos = address.rfind(keyword)
        if pos != -1:
            cut_position = pos + len(keyword)
            if cut_position > found_pos:
                found_pos = cut_position
                found_keyword = keyword

    if found_pos != -1 and found_keyword:
        address = address[:found_pos].strip()
        print(f"  ✅ Đã cắt tại từ khóa: '{found_keyword}'")

    return address


def read_checkpoint():
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, 'r') as f:
            try:
                content = f.read().strip()
                return int(content) if content else 0
            except (ValueError, IndexError):
                return 0
    return 0


def write_checkpoint(page_num):
    with open(checkpoint_file, 'w') as f:
        f.write(str(page_num))


# --- PHẦN CHÍNH: THU THẬP VÀ XỬ LÝ DỮ LIỆU ---
try:
    all_extracted_data = []
    all_raw_titles = []

    # --- Lấy tổng số trang một cách tự động ---
    print("Đang kiểm tra tổng số trang...")
    initial_payload = {
        "gdbsDivision": "440300",
        "gdbsOrgNum": "MB2C94128",
        "keywords": "电梯建设",
        "page": 1,
        "position": "title",
        "range": "site",
        "recommand": 1,
        "service_area": 755,
        "site_id": "755016",
        "sort": "smart"
    }

    try:
        response = requests.post(api_url, headers=headers, json=initial_payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        print(f"✅ Kết nối API thành công!")

        # FIXED: Correct data structure parsing
        if 'data' in data and 'news' in data['data']:
            results = data['data']['news']['list']
            total_items = data['data']['news'].get('total', 0)

            # Calculate total pages (assuming 20 items per page)
            total_pages = (total_items + 19) // 20  # Ceiling division

            print(f"📈 Tổng số kết quả: {total_items}")
            print(f"📄 Tổng số trang: {total_pages}")
            print(f"📝 Số mục trên trang 1: {len(results)}")

            if total_pages == 0:
                print("❌ Không tìm thấy kết quả nào với từ khóa '电梯建设'")
                exit()
        else:
            print("❌ Cấu trúc dữ liệu không đúng")
            print(f"📋 Các keys có trong data: {list(data.get('data', {}).keys())}")
            exit()

    except requests.exceptions.RequestException as e:
        print(f"❌ Lỗi kết nối API: {e}")
        exit()
    except json.JSONDecodeError as e:
        print(f"❌ Lỗi phân tích JSON: {e}")
        exit()

    # --- Bắt đầu quá trình crawl ---
    last_completed_page = read_checkpoint()
    start_page = last_completed_page + 1

    if start_page > total_pages:
        print(f"✅ Checkpoint cho thấy đã thu thập xong {last_completed_page}/{total_pages} trang.")
    else:
        if start_page > 1:
            print(f"🔄 Tiếp tục từ trang {start_page}...")

        for page_num in range(start_page, total_pages + 1):
            payload = initial_payload.copy()
            payload['page'] = page_num

            print(f"--- Đang thu thập trang {page_num}/{total_pages} ---")

            response = None
            for attempt in range(retry_attempts):
                try:
                    response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                    response.raise_for_status()
                    data = response.json()
                    break
                except requests.exceptions.RequestException as req_err:
                    print(f"  ❌ Lỗi kết nối (lần {attempt + 1}/{retry_attempts}): {req_err}")
                    if attempt < retry_attempts - 1:
                        time.sleep(retry_delay)
                    else:
                        print("  🚫 Đã hết số lần thử lại. Bỏ qua trang này.")
                        continue

            if response is None:
                continue

            try:
                # FIXED: Correct data structure
                results = data.get('data', {}).get('news', {}).get('list', [])
            except:
                print(f"  ❌ Không thể lấy dữ liệu từ trang {page_num}")
                continue

            if not results:
                print(f"ℹ️ Trang {page_num} không có dữ liệu.")
                break

            valid_items_count = 0
            for item in results:
                full_title = item.get('title', '')
                # Remove <em> tags from title
                full_title = re.sub(r'<.*?>', '', full_title)
                content = item.get('content', '')
                date_string = item.get('pub_time', '')
                url = item.get('url', '')

                # LỌC: Chỉ xử lý nếu tiêu đề chứa từ khóa địa chỉ
                if not any(keyword in full_title for keyword in ADDRESS_INDICATOR_KEYWORDS):
                    print(f"  ➖ Bỏ qua (không có từ khóa địa chỉ): {full_title[:50]}...")
                    continue

                # XỬ LÝ
                if full_title and date_string:
                    all_raw_titles.append(full_title)

                    # Trích xuất quận (ưu tiên tiêu đề trước)
                    district = get_district(full_title, content)

                    # Làm sạch tiêu đề và phân tích địa chỉ
                    cleaned_address = parse_title_hybrid_improved(full_title)
                    xiaoqu, dongshu, danyuan = parse_address_components(cleaned_address)

                    # Phân tích ngày
                    if 'T' in date_string:
                        date_part = date_string.split('T')[0]
                        parts = date_part.split('-')
                    else:
                        parts = date_string.split('-')

                    year, month, day = (parts[0], parts[1], parts[2]) if len(parts) >= 3 else ("", "", "")

                    all_extracted_data.append({
                        '区': district,
                        '小区': xiaoqu,
                        '栋数': dongshu,
                        '单元': danyuan,
                        '年': year,
                        '月': month,
                        '日': day
                    })

                    valid_items_count += 1
                    print(f"  ✅ [{valid_items_count}] {district} - {xiaoqu} {dongshu} {danyuan}")

            print(f"  📊 Trang {page_num}: {valid_items_count}/{len(results)} mục hợp lệ")
            write_checkpoint(page_num)
            time.sleep(request_delay)

        # --- Ghi dữ liệu ra file ---
        print(f"\n🎉 Thu thập hoàn tất. Tổng cộng {len(all_extracted_data)} tiêu đề hợp lệ.\n")

        if all_raw_titles:
            with open(txt_filename, 'w', encoding='utf-8') as f:
                f.write('\n'.join(all_raw_titles) + '\n')
            print(f"📄 Đã ghi {len(all_raw_titles)} tiêu đề vào file '{txt_filename}'")

        if all_extracted_data:
            columns_order = ['区', '小区', '栋数', '单元', '年', '月', '日']
            new_df = pd.DataFrame(all_extracted_data)[columns_order]

            try:
                if not os.path.exists(excel_filename):
                    pd.DataFrame(columns=columns_order).to_excel(excel_filename, sheet_name=sheet_name_to_update,
                                                                 index=False)
                    print(f"📁 Đã tạo file Excel mới: {excel_filename}")

                book = load_workbook(excel_filename)
                if sheet_name_to_update not in book.sheetnames:
                    book.create_sheet(sheet_name_to_update)
                sheet = book[sheet_name_to_update]

                last_data_row = find_last_row_with_data(sheet)
                start_row = last_data_row + 1

                if start_row <= 1:
                    for col_idx, header in enumerate(columns_order, 1):
                        sheet.cell(row=1, column=col_idx, value=header)
                    start_row = 2

                print(f"📝 Đang ghi {len(new_df)} dòng dữ liệu...")
                for i, row_data in new_df.iterrows():
                    current_row = start_row + i
                    for col_idx, col_name in enumerate(columns_order, 1):
                        sheet.cell(row=current_row, column=col_idx, value=row_data[col_name])

                book.save(excel_filename)
                print("✅ Hoàn tất! File Excel đã được cập nhật.")
            except Exception as ex:
                print(f"❌ Lỗi khi ghi file Excel: {ex}")
        else:
            print("ℹ️ Không có dữ liệu hợp lệ để ghi vào file Excel.")

except Exception as e:
    print(f"❌ Đã xảy ra lỗi: {e}")
    import traceback

    traceback.print_exc()